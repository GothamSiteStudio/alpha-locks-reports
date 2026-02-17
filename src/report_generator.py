"""Report generation for Alpha Locks Reports"""
import pandas as pd
from datetime import date
from typing import List, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import Job, JobResult, Technician
from .calculator import CommissionCalculator
from config import COMPANY_NAME, EXCEL_STYLES


class ReportGenerator:
    """
    Generates Excel reports for technician commissions.
    """
    
    def __init__(self, technician: Technician):
        self.technician = technician
        self.calculator = CommissionCalculator()
        self.results: List[JobResult] = []
    
    def add_jobs(self, jobs: List[Job]) -> None:
        """Add jobs and calculate results."""
        self.results.extend(self.calculator.calculate_batch(jobs))
    
    def clear(self) -> None:
        """Clear all results."""
        self.results = []
    
    def to_dataframe(self) -> pd.DataFrame:
        """
        Convert results to a pandas DataFrame.
        """
        data = []
        for r in self.results:
            data.append({
                'Date': r.job.job_date.strftime('%d/%m/%Y') if r.job.job_date else '',
                'Address': r.job.address,
                '%': 'Custom' if r.job.tech_amount is not None else f"{int(r.job.commission_rate * 100)}%",
                'Total': f"${r.job.total:,.2f}",
                'Parts': f"${r.job.parts:,.2f}" if r.job.parts > 0 else '',
                'Cash': f"${r.job.cash_amount:,.2f}" if r.job.cash_amount > 0 else '',
                'CC': f"${r.job.cc_amount:,.2f}" if r.job.cc_amount > 0 else '',
                'Check': f"${r.job.check_amount:,.2f}" if r.job.check_amount > 0 else '',
                'FEE': f"${r.job.fee:,.2f}" if r.job.fee > 0 else '',
                'Tech Profit': f"${r.tech_profit:,.2f}",
                'Balance': f"${r.balance:,.2f}"
            })
        
        return pd.DataFrame(data)
    
    def get_summary_row(self) -> dict:
        """Get summary row data."""
        summary = self.calculator.calculate_summary(self.results)
        return {
            'Date': f"{summary['job_count']} Jobs",
            'Address': '',
            '%': '',
            'Total': summary['total_sales'],
            'Parts': summary['total_parts'],
            'Cash': summary['total_cash'],
            'CC': summary['total_cc'],
            'Check': summary['total_check'],
            'FEE': summary['total_fee'],
            'Tech Profit': summary['total_tech_profit'],
            'Balance': summary['total_balance']
        }
    
    def get_date_range(self) -> tuple:
        """Get the date range of jobs."""
        dates = [r.job.job_date for r in self.results if r.job.job_date]
        if not dates:
            return None, None
        return min(dates), max(dates)
    
    def export_excel(self, filepath: str) -> None:
        """
        Export report to Excel file.
        
        Args:
            filepath: Path to save the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Commission Report"
        
        # Styles
        header_fill = PatternFill(start_color=EXCEL_STYLES['header_bg_color'],
                                   end_color=EXCEL_STYLES['header_bg_color'],
                                   fill_type='solid')
        summary_fill = PatternFill(start_color=EXCEL_STYLES['summary_bg_color'],
                                    end_color=EXCEL_STYLES['summary_bg_color'],
                                    fill_type='solid')
        header_font = Font(name=EXCEL_STYLES['font_name'], 
                          size=EXCEL_STYLES['font_size'], 
                          bold=True)
        title_font = Font(name=EXCEL_STYLES['font_name'], 
                         size=14, 
                         bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title section
        ws['A1'] = COMPANY_NAME
        ws['A1'].font = title_font
        
        ws['A2'] = f"Technician: {self.technician.name}"
        ws['A2'].font = Font(size=12, bold=True)
        
        start_date, end_date = self.get_date_range()
        if start_date and end_date:
            date_str = f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
            ws['A3'] = f"Period: {date_str}"
        
        # Data starts at row 5
        df = self.to_dataframe()
        start_row = 5
        
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=start_row + row_idx + 1, column=col_idx, value=value)
                cell.border = border
                if col_idx >= 4:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)) and value != '':
                        cell.number_format = '$#,##0.00'
        
        # Summary row
        summary_row = start_row + len(df) + 1
        summary_data = self.get_summary_row()
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=summary_row, column=col_idx, value=summary_data[col_name])
            cell.fill = summary_fill
            cell.font = Font(bold=True)
            cell.border = border
            if col_idx >= 4:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(summary_data[col_name], (int, float)):
                    cell.number_format = '$#,##0.00'
        
        # Adjust column widths
        column_widths = [12, 35, 8, 12, 12, 12, 12, 12, 10, 12, 12]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
        # Save
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
    
    def export_pdf(self, filepath: str) -> None:
        """Export report to PDF (future implementation)."""
        raise NotImplementedError("PDF export coming soon")
